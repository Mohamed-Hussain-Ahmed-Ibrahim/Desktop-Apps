import customtkinter as ctk
from customtkinter import CTkFrame, CTkLabel, CTkButton, CTkScrollableFrame, CTkTabview
import sqlite3
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import filedialog
from datetime import datetime
import numpy as np
import seaborn as sns
from tkinter import ttk
import pandas as pd
from tkinter.messagebox import showerror, showinfo
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module='matplotlib')
class ProstateStatisticsGUI(CTkFrame):
    def __init__(self, parent, db_path="DB/Informations.db", **kwargs):
        super().__init__(parent, **kwargs)
        self.parent = parent
        self.db_path = db_path
        self._after_ids = []  # Track after() callbacks
        
        # Handle window closing properly
        # Get the root window properly
        self.root = self._get_root_window()
        
        # Set up proper closing handling
        self._setup_window_protocols()
        
        # Track resources
        self._after_ids = []
        self._figures = []
        # Professional color scheme
        self.colors = {
            'primary': '#1e3a8a',      # Deep blue
            'secondary': '#3b82f6',     # Bright blue
            'accent': '#10b981',        # Emerald green
            'success': '#059669',       # Dark green
            'warning': '#f59e0b',       # Amber
            'danger': '#dc2626',        # Red
            'light': '#f8fafc',         # Light gray
            'dark': '#1e293b',          # Dark gray
            'white': '#ffffff',         # White
            'sidebar': '#0f172a',       # Dark sidebar
            'card': '#ffffff',          # Card background
            'border': '#e2e8f0',        # Border color
            'text_dark': '#1f2937',     # Dark text
            'text_light': '#6b7280',    # Light text
            'chart_colors': ['#3b82f6', '#10b981', '#f59e0b', '#dc2626', '#8b5cf6', '#06b6d4']
        }
        
        # Configure matplotlib style
        plt.style.use('seaborn-v0_8')
        sns.set_palette(self.colors['chart_colors'])
        # Initialize empty data structures first
        self.patient_data = pd.DataFrame()
        self.side_data = pd.DataFrame()
        self.bottles_data = pd.DataFrame()
        
        # Then load data before setting up GUI
        self.load_data()
        self.setup_gui()
    def _get_root(self):
        """Get the root window from any widget"""
        widget = self
        while widget.master is not None:
            widget = widget.master
        return widget
    def _get_root_window(self):
        """Safely get the root window regardless of widget hierarchy"""
        widget = self
        while hasattr(widget, 'master') and widget.master is not None:
            widget = widget.master
        return widget

    def _setup_window_protocols(self):
        """Set up window closing protocols"""
        if hasattr(self.root, 'protocol'):
            self.root.protocol("WM_DELETE_WINDOW", self.cleanup_and_quit)
        
        # For additional windows you might create:
        # if isinstance(self.parent, ctk.CTkToplevel):
        #     self.parent.protocol("WM_DELETE_WINDOW", self.cleanup_and_quit)

    def cleanup_and_quit(self):
        """Comprehensive cleanup before quitting"""
        # 1. Close all matplotlib figures
        plt.close('all')
        
        # 2. Cancel all pending after() callbacks
        for after_id in self._after_ids:
            try:
                self.after_cancel(after_id)
            except:
                pass
        
        # 3. Close database connections
        if hasattr(self, 'conn'):
            try:
                self.conn.close()
            except:
                pass
        
        # 4. Destroy the window
        try:
            if hasattr(self, 'root'):
                self.root.destroy()
            elif hasattr(self, 'parent'):
                self.parent.destroy()
        except:
            pass
        
        # 5. Force quit if needed (as last resort)
        import os
        os._exit(0)  # Nuclear option - only use if absolutely necessary

    def safe_after(self, ms, func, *args):
        """Wrapper for after() that tracks callbacks"""
        after_id = self.after(ms, func, *args)
        self._after_ids.append(after_id)
        return after_id

    def create_figure(self):
        """Create and track matplotlib figures"""
        fig = plt.figure()
        self._figures.append(fig)
        return fig
    def on_close(self):
        """Clean up before closing the application"""
        # Close all matplotlib figures
        import matplotlib.pyplot as plt
        plt.close('all')
        
        # Cancel any pending after() callbacks
        for after_id in self._after_ids:
            self.after_cancel(after_id)
        
        # Close database connection
        if hasattr(self, 'conn'):
            self.conn.close()
        
        # Destroy the window
        self.parent.destroy()
    def setup_gui(self):
        """Initialize the main GUI layout"""
        # Configure grid weights
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        # Create sidebar
        self.create_sidebar()
        
        # Create main content area with tabs
        self.create_main_content()
        
    def create_sidebar(self):
        """Create professional sidebar with navigation"""
        self.sidebar_frame = CTkFrame(self, fg_color=self.colors['sidebar'], 
                                     corner_radius=0, width=280)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        self.sidebar_frame.grid_rowconfigure(8, weight=1)
        
        # Header
        header_frame = CTkFrame(self.sidebar_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, pady=(20, 30), padx=20, sticky="ew")
        
        title_label = CTkLabel(header_frame, text="Prostate Analytics", 
                              font=("Arial Black", 20, "bold"), 
                              text_color=self.colors['white'])
        title_label.pack()
        
        subtitle_label = CTkLabel(header_frame, text="Medical Database Intelligence", 
                                 font=("Arial", 12), 
                                 text_color=self.colors['accent'])
        subtitle_label.pack(pady=(5, 0))
        
        # Navigation buttons
        nav_buttons = [
            ("📊 Overview", self.show_overview),
            ("👥 Patient Demographics", self.show_demographics),
            ("🔬 Clinical Analysis", self.show_clinical),
            ("📈 Trends & Patterns", self.show_trends),
            ("🎯 Risk Assessment", self.show_risk_analysis),
            ("📋 Detailed Reports", self.show_reports),
            ("⚙️ Settings", self.show_settings)
        ]
        
        for i, (text, command) in enumerate(nav_buttons):
            btn = CTkButton(self.sidebar_frame, text=text, 
                           fg_color="transparent", 
                           hover_color=self.colors['primary'],
                           corner_radius=10, height=40,
                           anchor="w", font=("Arial", 14),
                           text_color=self.colors['white'],
                           command=command)
            btn.grid(row=i+1, column=0, padx=20, pady=5, sticky="ew")
        
        # Status section
        status_frame = CTkFrame(self.sidebar_frame, fg_color=self.colors['dark'])
        status_frame.grid(row=8, column=0, padx=20, pady=20, sticky="sew")
        
        status_label = CTkLabel(status_frame, text="Database Status", 
                               font=("Arial Black", 14, "bold"), 
                               text_color=self.colors['white'])
        status_label.pack(pady=(10, 5))
        
        # Will be updated with actual data
        self.status_text = CTkLabel(status_frame, text="Loading...", 
                                   font=("Arial", 11), 
                                   text_color=self.colors['text_light'])
        self.status_text.pack(pady=(0, 10))
    def create_psa_distribution_figure(self):
        """Create PSA distribution figure for PDF"""
        fig, ax = plt.subplots(figsize=(8, 5))
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            psa_data = self.patient_data['TPSA'].dropna()
            ax.hist(psa_data, bins=20, color=self.colors['accent'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('TPSA (ng/ml)')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Total PSA Distribution')
            ax.grid(True, alpha=0.3)
        return fig
    def show_reports(self):
        """Display reports generation and export options"""
        # Clear current content
        for widget in self.reports_tab.winfo_children():
            widget.destroy()
            
        # Create scrollable frame for reports
        scroll_frame = CTkScrollableFrame(self.reports_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Reports header
        header_frame = CTkFrame(scroll_frame, fg_color=self.colors['primary'])
        header_frame.pack(fill="x", pady=(0, 20))
        
        header_label = CTkLabel(header_frame, text="Report Generation", 
                            font=("Arial Black", 18, "bold"), 
                            text_color=self.colors['white'])
        header_label.pack(pady=15)
        
        # Report types section
        types_frame = CTkFrame(scroll_frame, fg_color=self.colors['white'])
        types_frame.pack(fill="x", pady=(0, 20))
        
        types_title = CTkLabel(types_frame, text="Select Report Type", 
                            font=("Arial Black", 14, "bold"), 
                            text_color=self.colors['primary'])
        types_title.pack(pady=(15, 5), anchor="w", padx=15)
        
        # Report type options
        report_types = [
            ("📊 Summary Report", "Comprehensive overview of all statistics"),
            ("👥 Demographic Report", "Detailed demographic analysis"),
            ("🔬 Clinical Findings", "Clinical data and test results"),
            ("📈 Trend Analysis", "Temporal trends and patterns"),
            ("🎯 Risk Assessment", "Patient risk stratification")
        ]
        
        self.report_type_var = ctk.StringVar(value=report_types[0][0])
        
        for text, desc in report_types:
            report_frame = CTkFrame(types_frame, fg_color="transparent")
            report_frame.pack(fill="x", padx=15, pady=5)
            
            rb = ctk.CTkRadioButton(report_frame, text=text, 
                                variable=self.report_type_var, 
                                value=text,
                                font=("Arial", 14))
            rb.pack(side="left")
            
            desc_label = CTkLabel(report_frame, text=desc,
                                text_color=self.colors['text_light'],
                                font=("Arial", 12))
            desc_label.pack(side="left", padx=10)
        
        # Export options section
        export_frame = CTkFrame(scroll_frame, fg_color=self.colors['white'])
        export_frame.pack(fill="x", pady=(0, 20))
        
        export_title = CTkLabel(export_frame, text="Export Options", 
                            font=("Arial Black", 14, "bold"), 
                            text_color=self.colors['primary'])
        export_title.pack(pady=(15, 5), anchor="w", padx=15)
        
        # Format selection
        format_frame = CTkFrame(export_frame, fg_color="transparent")
        format_frame.pack(fill="x", padx=15, pady=5)
        
        format_label = CTkLabel(format_frame, text="Format:", 
                            font=("Arial", 12), 
                            text_color=self.colors['text_dark'])
        format_label.pack(side="left")
        
        self.format_var = ctk.StringVar(value="PDF")
        format_options = ["PDF", "Excel", "CSV", "HTML"]
        format_menu = ctk.CTkOptionMenu(format_frame, values=format_options,
                                    variable=self.format_var)
        format_menu.pack(side="left", padx=10)
        
        # File name entry
        name_frame = CTkFrame(export_frame, fg_color="transparent")
        name_frame.pack(fill="x", padx=15, pady=5)
        
        name_label = CTkLabel(name_frame, text="File Name:", 
                            font=("Arial", 12), 
                            text_color=self.colors['text_dark'])
        name_label.pack(side="left")
        
        default_name = f"prostate_report_{datetime.now().strftime('%Y%m%d')}"
        self.filename_entry = ctk.CTkEntry(name_frame, 
                                        placeholder_text=default_name,
                                        width=300)
        self.filename_entry.pack(side="left", padx=10)
        
        # Generate and export buttons
        button_frame = CTkFrame(scroll_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(0, 20))
        
        preview_btn = CTkButton(button_frame, text="Preview Report",
                            fg_color=self.colors['secondary'],
                            command=self.preview_report)
        preview_btn.pack(side="left", padx=10)
        
        export_btn = CTkButton(button_frame, text="Export Report",
                            fg_color=self.colors['accent'],
                            command=self.export_report)
        export_btn.pack(side="left", padx=10)
        
        # Add reports tab if it doesn't exist
        if "Reports" not in self.tabview._tab_dict:
            self.reports_tab = self.tabview.add("Reports")
        
        # Switch to reports tab
        self.tabview.set("Reports")
    def configure_unicode_font(self, pdf):
        """Configure a Unicode-compatible font"""
        try:
            pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
            pdf.set_font("DejaVu", "", 12)
        except:
            pdf.set_font("Arial", "", 12)

    def safe_text(self, text):
        """Ensure text is PDF-compatible"""
        return str(text).encode('latin-1', 'replace').decode('latin-1')
    def _get_summary_metrics(self):
        """Calculate and return summary metrics"""
        metrics = []
        
        if not self.patient_data.empty:
            # Total patients
            metrics.append(f"Total Patients: {len(self.patient_data)}")
            
            # Average age
            if 'Age' in self.patient_data.columns:
                avg_age = self.patient_data['Age'].mean()
                metrics.append(f"Average Age: {avg_age:.1f} years")
            
            # High PSA count
            if 'TPSA' in self.patient_data.columns:
                high_psa = sum(self.patient_data['TPSA'] > 4)
                metrics.append(f"Patients with PSA >4: {high_psa}")
            
            # Family history
            if 'Family' in self.patient_data.columns:
                family_pos = sum(self.patient_data['Family'] == '+ve')
                metrics.append(f"Family History Positive: {family_pos}")
        
        return metrics

    def _create_age_chart(self):
        """Create age distribution chart"""
        fig, ax = plt.subplots(figsize=(8, 5))
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            ages = self.patient_data['Age'].dropna()
            ax.hist(ages, bins=15, color='#1e3a8a', alpha=0.7, edgecolor='white')
            ax.set_xlabel('Age')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Patient Age Distribution')
            ax.grid(True, alpha=0.3)
        return fig

    def _create_psa_chart(self):
        """Create PSA distribution chart"""
        fig, ax = plt.subplots(figsize=(8, 5))
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            psa_data = self.patient_data['TPSA'].dropna()
            ax.hist(psa_data, bins=20, color='#3b82f6', alpha=0.7, edgecolor='white')
            ax.set_xlabel('TPSA (ng/ml)')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Total PSA Distribution')
            ax.grid(True, alpha=0.3)
        return fig

    def _add_figure_to_pdf(self, pdf, title, fig):
        """Add matplotlib figure to PDF"""
        import tempfile
        import os
        
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_chart_{title.replace(' ', '_')}.png")
        
        try:
            fig.savefig(temp_file, dpi=300, bbox_inches='tight')
            pdf.ln(10)
            pdf.set_font("", "B", 12)
            pdf.cell(200, 10, txt=title, ln=1)
            pdf.set_font("", "", 12)
            pdf.image(temp_file, x=10, w=180)
        finally:
            if os.path.exists(temp_file):
                os.remove(temp_file)
    def add_summary_metrics_to_pdf(self, pdf):
        """Add summary content to PDF"""
        # Add metrics
        pdf.set_font("", "B", 14)
        pdf.cell(200, 10, txt="Key Statistics", ln=1)
        pdf.set_font("", "", 12)
        
        metrics = [
            f"Total Patients: {len(self.patient_data)}",
            f"Average Age: {self.patient_data['Age'].mean():.1f}",
            f"High PSA Count: {sum(self.patient_data['TPSA'] > 4)}"
        ]
        
        for metric in metrics:
            pdf.cell(200, 10, txt=metric, ln=1)
        
        # Add charts
        if not self.patient_data.empty:
            if 'Age' in self.patient_data.columns:
                fig = self._create_age_chart()
                self._add_figure_to_pdf(pdf, "Age Distribution", fig)
                plt.close(fig)
            
            if 'TPSA' in self.patient_data.columns:
                fig = self._create_psa_chart()
                self._add_figure_to_pdf(pdf, "PSA Distribution", fig)
                plt.close(fig)
    def export_to_pdf(self, report_type, filename):
        """Export report to PDF format"""
        try:
            from fpdf import FPDF
            import tempfile
            import os
            
            # Create PDF object
            pdf = FPDF()
            pdf.add_page()
            
            # Configure Unicode font
            self.configure_unicode_font(pdf)
            
            # Add title
            pdf.set_font_size(16)
            pdf.cell(200, 10, txt=self.safe_text(report_type), ln=1, align='C')
            pdf.set_font_size(12)
            
            # Add content based on report type
            if "Summary" in report_type:
                self.add_summary_to_pdf(pdf)
            elif "Demographic" in report_type:
                self.add_demographics_to_pdf(pdf)
            elif "Clinical" in report_type:
                self.add_clinical_to_pdf(pdf)  # Fixed spelling here
            elif "Trend" in report_type:
                self.add_trends_to_pdf(pdf)  # This will now work
            elif "Risk" in report_type:
                self.add_risk_to_pdf(pdf)
            
            # Save the PDF
            pdf.output(filename)
            showinfo(title="Export Successful",
                    message=f"{report_type} exported successfully")
                    
        except ImportError:
            showerror(title="Export Failed",
                    message="FPDF library is required for PDF export. Install with: pip install fpdf2")
            plt.close('all')
        except Exception as e:
            showerror(title="Export Failed",
                    message=f"Error exporting report: {str(e)}")
            plt.close('all')

    def add_summary_to_pdf(self, pdf):
        """Add summary report content to PDF"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Key Statistics", ln=1)
        pdf.set_font("Arial", size=12)
        
        # Calculate and add metrics
        total_patients = len(self.patient_data) if not self.patient_data.empty else 0
        avg_age = self.patient_data['Age'].mean() if not self.patient_data.empty and 'Age' in self.patient_data.columns else 0
        high_psa = len(self.patient_data[self.patient_data['TPSA'] > 4]) if not self.patient_data.empty and 'TPSA' in self.patient_data.columns else 0
        
        metrics = [
            f"Total Patients: {total_patients}",
            f"Average Age: {avg_age:.1f} years",
            f"Patients with PSA >4: {high_psa}",
            f"Family History Positive: {len(self.patient_data[self.patient_data['Family'] == '+ve']) if not self.patient_data.empty and 'Family' in self.patient_data.columns else 0}"
        ]
        
        for metric in metrics:
            pdf.cell(200, 10, txt=metric, ln=1)
        
        # Add charts using the correct figure method
        if not self.patient_data.empty:
            if 'Age' in self.patient_data.columns:
                fig = self.create_age_distribution_figure()
                self.add_figure_to_pdf(pdf, "Age Distribution", fig)
                plt.close(fig)
            
            if 'TPSA' in self.patient_data.columns:
                fig = self.create_psa_distribution_figure()
                self.add_figure_to_pdf(pdf, "PSA Distribution", fig)
                plt.close(fig)

    def create_age_distribution_figure(self):
        """Create age distribution figure for PDF"""
        fig, ax = plt.subplots(figsize=(8, 5))
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            ages = self.patient_data['Age'].dropna()
            ax.hist(ages, bins=15, color=self.colors['primary'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('Age')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Patient Age Distribution')
            ax.grid(True, alpha=0.3)
        return fig

    def create_psa_distribution_figure(self):
        """Create PSA distribution figure for PDF"""
        fig, ax = plt.subplots(figsize=(8, 5))
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            psa_data = self.patient_data['TPSA'].dropna()
            ax.hist(psa_data, bins=20, color=self.colors['accent'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('TPSA (ng/ml)')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Total PSA Distribution')
            ax.grid(True, alpha=0.3)
        return fig
    def add_chart_to_pdf(self, pdf, title, fig):
        """Add matplotlib figure to PDF"""
        import tempfile
        import os
        
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_chart_{title.replace(' ', '_')}.png")
        
        try:
            fig.savefig(temp_file, dpi=300, bbox_inches='tight')
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt=title, ln=1)
            pdf.set_font("Arial", size=12)
            pdf.image(temp_file, x=10, w=180)
        finally:
            if os.path.exists(temp_file):
                os.remove(temp_file)
    def on_preview_close(self, window):
        window.destroy()
    def preview_report(self):
        """Generate a preview of the selected report"""
        report_type = self.report_type_var.get()
        
        # Create preview window
        preview_window = ctk.CTkToplevel(self)
        preview_window.title(f"Report Preview - {report_type}")
        preview_window.geometry("800x600")
        preview_window.protocol("WM_DELETE_WINDOW", lambda: self.on_preview_close(preview_window))
        # Add content based on report type
        scroll_frame = CTkScrollableFrame(preview_window)
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        title_label = CTkLabel(scroll_frame, text=report_type,
                            font=("Arial Black", 16, "bold"),
                            text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        # Add sample content based on report type
        if "Summary" in report_type:
            self.add_summary_preview(scroll_frame)
        elif "Demographic" in report_type:
            self.add_demographic_preview(scroll_frame)
        elif "Clinical" in report_type:
            self.add_clinical_preview(scroll_frame)
        elif "Trend" in report_type:
            self.add_trend_preview(scroll_frame)
        elif "Risk" in report_type:
            self.add_risk_preview(scroll_frame)
        
        close_btn = CTkButton(preview_window, text="Close",
                            command=preview_window.destroy)
        close_btn.pack(pady=10)

    def add_summary_preview(self, parent):
        """Add summary report preview content"""
        # Key metrics
        metrics_frame = CTkFrame(parent, fg_color=self.colors['light'])
        metrics_frame.pack(fill="x", pady=10)
        
        metrics_title = CTkLabel(metrics_frame, text="Key Metrics",
                            font=("Arial Black", 14),
                            text_color=self.colors['primary'])
        metrics_title.pack(pady=5)
        
        # Calculate metrics (simplified for preview)
        total_patients = len(self.patient_data) if not self.patient_data.empty else 0
        avg_age = self.patient_data['Age'].mean() if not self.patient_data.empty and 'Age' in self.patient_data.columns else 0
        high_psa = len(self.patient_data[self.patient_data['TPSA'] > 4]) if not self.patient_data.empty and 'TPSA' in self.patient_data.columns else 0
        
        metrics_text = f"""
        • Total Patients: {total_patients}
        • Average Age: {avg_age:.1f} years
        • Patients with PSA >4: {high_psa}
        """
        
        metrics_content = CTkLabel(metrics_frame, text=metrics_text,
                                font=("Arial", 12),
                                text_color=self.colors['text_dark'])
        metrics_content.pack(pady=5)
        
        # Add sample charts
        self.add_preview_chart(parent, "Age Distribution", "age_distribution.png")
        self.add_preview_chart(parent, "PSA Distribution", "psa_distribution.png")

    def add_demographic_preview(self, parent):
        """Add demographic report preview content"""
        # Demographic summary
        summary_frame = CTkFrame(parent, fg_color=self.colors['light'])
        summary_frame.pack(fill="x", pady=10)
        
        summary_title = CTkLabel(summary_frame, text="Demographic Summary",
                            font=("Arial Black", 14),
                            text_color=self.colors['primary'])
        summary_title.pack(pady=5)
        
        # Add sample charts
        self.add_preview_chart(parent, "Age Groups", "age_groups.png")
        self.add_preview_chart(parent, "Nationality Distribution", "nationality.png")

    def add_demographics_to_pdf(self, pdf):
        """Add demographic report content to PDF"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Demographic Statistics", ln=1)
        pdf.set_font("Arial", size=12)
        
        # Add age distribution
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="Age Distribution", ln=1)
            fig = self.create_age_distribution_figure()
            self.add_figure_to_pdf(pdf, "Age Distribution", fig)
            plt.close(fig)
        
        # Add nationality distribution if available
        if not self.patient_data.empty and 'Nationality' in self.patient_data.columns:
            nat_counts = self.patient_data['Nationality'].value_counts()
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="Nationality Distribution", ln=1)
            
            for nat, count in nat_counts.items():
                pdf.cell(200, 10, txt=f"{nat}: {count} patients ({count/len(self.patient_data):.1%})", ln=1)
    def add_preview_chart(self, parent, title, sample_image):
        """Add a sample chart to the preview"""
        chart_frame = CTkFrame(parent, fg_color=self.colors['white'])
        chart_frame.pack(fill="x", pady=10)
        
        chart_title = CTkLabel(chart_frame, text=title,
                            font=("Arial Black", 12),
                            text_color=self.colors['primary'])
        chart_title.pack(pady=5)
        
        # Placeholder for actual chart - in preview we'll just show a label
        chart_label = CTkLabel(chart_frame, text=f"[Chart: {title}]",
                            text_color=self.colors['text_light'])
        chart_label.pack(pady=20)
        
        # In a real implementation, you would generate the actual chart here

    def export_report(self):
        """Export the selected report to the chosen format"""
        from tkinter import filedialog
        
        report_type = self.report_type_var.get()
        file_format = self.format_var.get()
        default_name = f"prostate_report_{datetime.now().strftime('%Y%m%d')}"
        filename = self.filename_entry.get() or default_name
        
        # Add appropriate extension
        if file_format == "PDF":
            filename += ".pdf"
        elif file_format == "Excel":
            filename += ".xlsx"
        elif file_format == "CSV":
            filename += ".csv"
        elif file_format == "HTML":
            filename += ".html"
        
        # Ask user for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=f".{file_format.lower()}",
            initialfile=filename,
            filetypes=[(f"{file_format} files", f"*.{file_format.lower()}")]
        )
        
        if not file_path:  # User cancelled
            return
        
        try:
            if file_format == "PDF":
                self.export_to_pdf(report_type, file_path)
            elif file_format == "Excel":
                self.export_to_excel(report_type, file_path)
            elif file_format == "CSV":
                self.export_to_csv(report_type, file_path)
            elif file_format == "HTML":
                self.export_to_html(report_type, file_path)
            
            showinfo(title="Export Successful",message=f"{report_type} exported successfully")
                    
        except Exception as e:
            showerror(title="Export Failed",
                    message=f"Error exporting report: {str(e)}")

    def add_summary_to_pdf(self, pdf):
        """Add summary report content to PDF"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Key Statistics", ln=1)
        pdf.set_font("Arial", size=12)
        
        # Calculate and add metrics
        total_patients = len(self.patient_data) if not self.patient_data.empty else 0
        avg_age = self.patient_data['Age'].mean() if not self.patient_data.empty and 'Age' in self.patient_data.columns else 0
        high_psa = len(self.patient_data[self.patient_data['TPSA'] > 4]) if not self.patient_data.empty and 'TPSA' in self.patient_data.columns else 0
        
        metrics = [
            f"Total Patients: {total_patients}",
            f"Average Age: {avg_age:.1f} years",
            f"Patients with PSA >4: {high_psa}",
            f"Family History Positive: {len(self.patient_data[self.patient_data['Family'] == '+ve']) if not self.patient_data.empty and 'Family' in self.patient_data.columns else 0}"
        ]
        
        for metric in metrics:
            pdf.cell(200, 10, txt=metric, ln=1)
        
        # Add charts using the correct figure method
        if not self.patient_data.empty:
            if 'Age' in self.patient_data.columns:
                fig = self.create_age_distribution_figure()
                self.add_figure_to_pdf(pdf, "Age Distribution", fig)
                plt.close(fig)
            
            if 'TPSA' in self.patient_data.columns:
                fig = self.create_psa_distribution_figure()
                self.add_figure_to_pdf(pdf, "PSA Distribution", fig)
                plt.close(fig)
    def add_figure_to_pdf(self, pdf, title, fig):
        """Safely add a matplotlib figure to PDF"""
        import tempfile
        import os
        try:
            # Create temp file
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, f"temp_chart_{title.replace(' ', '_')}.png")
            
            # Save figure and add to PDF
            fig.savefig(temp_path, dpi=150, bbox_inches='tight', facecolor='white')
            pdf.ln(8)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 8, txt=title, ln=1)
            pdf.image(temp_path, x=10, w=190)
            
        except Exception as e:
            showerror("Chart Error", f"Could not add chart '{title}': {str(e)}")
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            plt.close(fig)
    def add_chart_to_pdf(self, pdf, title, fig):
        """Add a matplotlib figure to the PDF"""
        import tempfile
        import os
        
        # Save figure to temp file
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_chart_{title.replace(' ', '_')}.png")
        
        try:
            fig.savefig(temp_file, dpi=300, bbox_inches='tight')
            plt.close(fig)
            
            # Add to PDF
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt=title, ln=1)
            pdf.image(temp_file, x=10, w=180)
        finally:
            # Clean up
            if os.path.exists(temp_file):
                os.remove(temp_file)
                
        plt.close(fig)
    def _generate_demographics_html(self):
        """Generate HTML content for demographic report"""
        html = "<h2>Demographic Statistics</h2>\n"
        
        # Age distribution
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            html += "<h3>Age Distribution</h3>\n"
            ages = self.patient_data['Age'].dropna()
            age_groups = pd.cut(ages, bins=[0, 50, 60, 70, 80, 100], 
                            labels=['<50', '50-59', '60-69', '70-79', '≥80'])
            age_counts = age_groups.value_counts()
            
            html += "<table>\n<tr><th>Age Group</th><th>Count</th><th>Percentage</th></tr>\n"
            for group, count in age_counts.items():
                percent = (count / len(self.patient_data)) * 100
                html += f"<tr><td>{group}</td><td>{count}</td><td>{percent:.1f}%</td></tr>\n"
            html += "</table>\n"
        
        # Nationality distribution
        if not self.patient_data.empty and 'Nationality' in self.patient_data.columns:
            html += "<h3>Nationality Distribution</h3>\n"
            nat_counts = self.patient_data['Nationality'].value_counts()
            
            html += "<table>\n<tr><th>Nationality</th><th>Count</th><th>Percentage</th></tr>\n"
            for nat, count in nat_counts.items():
                percent = (count / len(self.patient_data)) * 100
                html += f"<tr><td>{nat}</td><td>{count}</td><td>{percent:.1f}%</td></tr>\n"
            html += "</table>\n"
        
        # Family history
        if not self.patient_data.empty and 'Family' in self.patient_data.columns:
            html += "<h3>Family History</h3>\n"
            family_counts = self.patient_data['Family'].value_counts()
            
            html += "<table>\n<tr><th>Family History</th><th>Count</th><th>Percentage</th></tr>\n"
            for status, count in family_counts.items():
                percent = (count / len(self.patient_data)) * 100
                html += f"<tr><td>{status}</td><td>{count}</td><td>{percent:.1f}%</td></tr>\n"
            html += "</table>\n"
        
        return html
    def _generate_trends_html(self):
        """Generate HTML content for trends analysis report"""
        html = "<h2>Trends Analysis</h2>\n"
        
        # Temporal Analysis
        if not self.patient_data.empty and 'Date' in self.patient_data.columns:
            try:
                dates = pd.to_datetime(self.patient_data['Date'], errors='coerce').dropna()
                if len(dates) > 0:
                    monthly_counts = dates.groupby(dates.dt.to_period('M')).size()
                    
                    html += "<h3>Patient Registration Over Time</h3>\n"
                    html += "<table>\n<tr><th>Month</th><th>Patient Count</th></tr>\n"
                    for month, count in monthly_counts.items():
                        html += f"<tr><td>{month}</td><td>{count}</td></tr>\n"
                    html += "</table>\n"
                    
                    # Add simple statistics
                    html += f"<p>Average monthly registrations: {monthly_counts.mean():.1f}</p>\n"
                    html += f"<p>Peak month: {monthly_counts.idxmax()} ({monthly_counts.max()} patients)</p>\n"
            except Exception as e:
                html += f"<p class='error'>Error processing dates: {str(e)}</p>\n"
        
        # PSA vs Age Analysis
        if (not self.patient_data.empty and 
            'Age' in self.patient_data.columns and 
            'TPSA' in self.patient_data.columns):
            
            data = self.patient_data[['Age', 'TPSA']].dropna()
            if len(data) > 0:
                html += "<h3>PSA vs Age Relationship</h3>\n"
                html += f"<p>Number of data points: {len(data)}</p>\n"
                
                if len(data) > 2:
                    z = np.polyfit(data['Age'], data['TPSA'], 1)
                    html += f"<p>Trend line equation: PSA = {z[0]:.2f}*Age + {z[1]:.2f}</p>\n"
                    html += f"<p>Correlation coefficient: {data['Age'].corr(data['TPSA']):.2f}</p>\n"
                
                # Create a simple table of statistics by age group
                data['AgeGroup'] = pd.cut(data['Age'], bins=[0, 50, 60, 70, 80, 100])
                stats = data.groupby('AgeGroup')['TPSA'].agg(['mean', 'count'])
                html += "<h4>Average PSA by Age Group</h4>\n"
                html += "<table>\n<tr><th>Age Group</th><th>Average PSA</th><th>Count</th></tr>\n"
                for group, row in stats.iterrows():
                    html += f"<tr><td>{group}</td><td>{row['mean']:.1f}</td><td>{row['count']}</td></tr>\n"
                html += "</table>\n"
        
        # Correlation Analysis
        numeric_cols = ['Age', 'TPSA', 'FPSA', 'PSA', 'WG_Volume_cc', 'A_Volume_cc']
        available_cols = [col for col in numeric_cols if col in self.patient_data.columns]
        
        if len(available_cols) >= 2:
            html += "<h3>Clinical Parameter Correlations</h3>\n"
            corr_matrix = self.patient_data[available_cols].corr()
            
            html += "<table>\n<tr><th>Parameter</th>"
            for col in available_cols:
                html += f"<th>{col}</th>"
            html += "</tr>\n"
            
            for i, row in corr_matrix.iterrows():
                html += f"<tr><td>{i}</td>"
                for col in available_cols:
                    html += f"<td>{row[col]:.2f}</td>"
                html += "</tr>\n"
            html += "</table>\n"
        
        return html
    def create_age_distribution_figure(self):
        """Create age distribution figure for PDF"""
        fig, ax = plt.subplots(figsize=(8, 5))
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            ages = self.patient_data['Age'].dropna()
            ax.hist(ages, bins=15, color=self.colors['primary'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('Age')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Patient Age Distribution')
            ax.grid(True, alpha=0.3)
        return fig
    def _generate_clinical_html(self):
        """Generate HTML content for clinical report"""
        html = "<h2>Clinical Statistics</h2>\n"
        
        # PSA Distribution
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            html += "<h3>PSA Distribution</h3>\n"
            psa_data = self.patient_data['TPSA'].dropna()
            html += f"<p>Average PSA: {psa_data.mean():.2f} ng/ml</p>\n"
            html += f"<p>Patients with PSA >4: {len(psa_data[psa_data > 4])}</p>\n"
            html += f"<p>Patients with PSA >10: {len(psa_data[psa_data > 10])}</p>\n"
        
        # DRE Results
        if not self.patient_data.empty and 'DRE' in self.patient_data.columns:
            html += "<h3>Digital Rectal Exam (DRE) Results</h3>\n"
            dre_counts = self.patient_data['DRE'].value_counts()
            
            html += "<table>\n<tr><th>DRE Result</th><th>Count</th><th>Percentage</th></tr>\n"
            for result, count in dre_counts.items():
                percent = (count / len(self.patient_data)) * 100
                html += f"<tr><td>{result}</td><td>{count}</td><td>{percent:.1f}%</td></tr>\n"
            html += "</table>\n"
        
        # MRI Results
        if not self.patient_data.empty and 'Mri_Result' in self.patient_data.columns:
            html += "<h3>MRI Results (PIRADS)</h3>\n"
            mri_counts = self.patient_data['Mri_Result'].value_counts()
            
            html += "<table>\n<tr><th>PIRADS Score</th><th>Count</th><th>Percentage</th></tr>\n"
            for result, count in mri_counts.items():
                percent = (count / len(self.patient_data)) * 100
                html += f"<tr><td>{result}</td><td>{count}</td><td>{percent:.1f}%</td></tr>\n"
            html += "</table>\n"
        
        # Volume Data
        if not self.patient_data.empty and 'WG_Volume_cc' in self.patient_data.columns:
            html += "<h3>Prostate Volume</h3>\n"
            volume_data = self.patient_data['WG_Volume_cc'].dropna()
            html += f"<p>Average Volume: {volume_data.mean():.1f} cc</p>\n"
            html += f"<p>Minimum Volume: {volume_data.min():.1f} cc</p>\n"
            html += f"<p>Maximum Volume: {volume_data.max():.1f} cc</p>\n"
        
        return html
    def export_to_excel(self, report_type, filename):
        """Export report to Excel format using openpyxl"""
        try:
            import openpyxl
            import pandas as pd
        
            # Use openpyxl as the engine
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                if "Summary" in report_type:
                    self.add_summary_to_excel(writer)
                # Add other report types as needed
                elif "Demographic" in report_type:
                    self.add_demographics_to_excel(writer)
                elif "Clinical" in report_type:
                    self.add_clinical_to_excel(writer)
                elif "Trend" in report_type:
                    self.add_trends_to_excel(writer)
                elif "Risk" in report_type:
                    self.add_risk_to_excel(writer)
            showinfo(title="Export Successful",message=f"{report_type} exported successfully to {filename}")
                
        except ImportError:
            showerror(title="Export Failed",
                    message="openpyxl library is required for Excel export. Install with: pip install openpyxl")
        except Exception as e:
            showerror(title="Export Failed",
                    message=f"Error exporting to Excel: {str(e)}") 
        # Save the Excel file
        writer.close()

    def export_to_excel(self, report_type, filename):
        """Export report to Excel format with proper formatting"""
        try:
            import pandas as pd
            import xlsxwriter
            
            # Create a Pandas Excel writer using xlsxwriter as the engine
            writer = pd.ExcelWriter(filename, engine='xlsxwriter')
            
            if "Summary" in report_type:
                # Create summary data
                summary_data = {
                    'Metric': ['Total Patients', 'Average Age', 'High PSA Count'],
                    'Value': [
                        len(self.patient_data),
                        self.patient_data['Age'].mean() if not self.patient_data.empty and 'Age' in self.patient_data.columns else 0,
                        len(self.patient_data[self.patient_data['TPSA'] > 4]) if not self.patient_data.empty and 'TPSA' in self.patient_data.columns else 0
                    ]
                }
                
                df = pd.DataFrame(summary_data)
                df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Access the xlsxwriter objects from the writer
                workbook = writer.book
                worksheet = writer.sheets['Summary']
                
                # Create a format object (THIS IS THE CORRECT WAY)
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'top',
                    'fg_color': '#D7E4BC',
                    'border': 1
                })
                
                # Apply the format to header row
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                
                # Auto-adjust column widths
                for col_num, value in enumerate(df.columns.values):
                    max_len = max(df[value].astype(str).map(len).max(), len(value)) + 2
                    worksheet.set_column(col_num, col_num, max_len)
                
            # Add other report types as needed...
            
            # Close the Pandas Excel writer and output the Excel file
            writer.close()
            
            showinfo(title="Export Successful",
                    message=f"{report_type} exported successfully to {filename}")
                    
        except ImportError:
            showerror(title="Export Failed",
                    message="xlsxwriter library is required for Excel export. Install with: pip install xlsxwriter")
        except Exception as e:
            showerror(title="Export Failed",
                    message=f"Error exporting to Excel: {str(e)}")
    def add_chart_to_pdf(self, pdf, title, parent_frame=None):
        """Add a chart to the PDF by creating it in a temporary frame"""
        import tempfile
        import os
        
        # Create a temporary frame if not provided
        if parent_frame is None:
            parent_frame = CTkFrame(self)
        
        # Clear the frame and create the chart
        for widget in parent_frame.winfo_children():
            widget.destroy()
        
        if title == "Age Distribution":
            self.create_age_distribution_chart(parent_frame)
        elif title == "PSA Distribution":
            self.create_psa_distribution_chart(parent_frame)
        # Add other chart types as needed
        
        # Save figure to temp file
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_chart_{title.replace(' ', '_')}.png")
        
        # Get the figure from the canvas
        for widget in parent_frame.winfo_children():
            if hasattr(widget, 'figure'):
                widget.figure.savefig(temp_file, dpi=300, bbox_inches='tight')
                plt.close(widget.figure)
                break
        
        # Add to PDF
        pdf.ln(10)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(200, 10, txt=title, ln=1)
        pdf.image(temp_file, x=10, w=180)
        
        # Clean up
        os.remove(temp_file)
    def export_to_csv(self, report_type, filename):
        """Export report to CSV format"""
        if "Summary" in report_type:
            data = {
                'Metric': ['Total Patients', 'Average Age', 'High PSA Count'],
                'Value': [
                    len(self.patient_data),
                    self.patient_data['Age'].mean() if not self.patient_data.empty and 'Age' in self.patient_data.columns else 0,
                    len(self.patient_data[self.patient_data['TPSA'] > 4]) if not self.patient_data.empty and 'TPSA' in self.patient_data.columns else 0
                ]
            }
            df = pd.DataFrame(data)
            df.to_csv(filename, index=False)


    def export_to_html(self, report_type, filename):
        """Export report to HTML format with proper Unicode handling"""
        try:
            html_content = self.generate_html_content(report_type)
            
            # Write with explicit UTF-8 encoding
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
        except UnicodeEncodeError as e:
            showerror("Encoding Error",
                    f"Cannot save special characters: {str(e)}\n"
                    "Try removing emojis or special symbols from the report.")
        except Exception as e:
            showerror("Export Error",
                    f"Failed to export HTML report: {str(e)}")

    def generate_html_content(self, report_type):
        """Generate HTML content with proper meta tags"""
        html_template = """<!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>{title}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ color: #1e3a8a; text-align: center; }}
            p {{ line-height: 1.6; color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
            th {{ background-color: #f2f2f2; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            .error {{ color: #dc2626; }}
            .highlight {{ background-color: #fef08a; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <p>Generated on: {date}</p>
        {content}
    </body>
    </html>"""
        
        # Generate report-specific content
        if "Summary" in report_type:
            content = self._generate_summary_html()
        elif "Demographic" in report_type:
            content = self._generate_demographics_html()
        elif "Clinical" in report_type:
            content = self._generate_clinical_html()
        elif "Trend" in report_type:
            content = self._generate_trends_html()  # Now properly implemented
        elif "Risk" in report_type:
            content = self._generate_risk_html()
        else:
            content = "<p>No content available for this report type</p>"
        
        return html_template.format(
            title=report_type,
            date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            content=content
            )
    def _generate_risk_html(self):
        """Generate professional HTML content for risk assessment report"""
        html = """
        <div class="risk-report">
            <h2>Prostate Cancer Risk Assessment</h2>
            <p class="report-date">Generated on: {date}</p>
            
            <div class="executive-summary">
                <h3>Executive Summary</h3>
        """.format(date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Calculate key risk metrics
        if not self.patient_data.empty:
            total_patients = len(self.patient_data)
            
            # PSA Risk Categories
            if 'TPSA' in self.patient_data.columns:
                psa_data = self.patient_data['TPSA'].dropna()
                low_risk = len(psa_data[psa_data <= 4])
                moderate_risk = len(psa_data[(psa_data > 4) & (psa_data <= 10)])
                high_risk = len(psa_data[psa_data > 10])
                
                html += f"""
                <div class="metric-card">
                    <h4>PSA Risk Stratification</h4>
                    <ul>
                        <li>Low Risk (≤4 ng/ml): {low_risk} patients ({low_risk/total_patients:.1%})</li>
                        <li>Moderate Risk (4-10 ng/ml): {moderate_risk} patients ({moderate_risk/total_patients:.1%})</li>
                        <li>High Risk (>10 ng/ml): {high_risk} patients ({high_risk/total_patients:.1%})</li>
                    </ul>
                </div>
                """

            # DRE Findings
            if 'DRE' in self.patient_data.columns:
                dre_counts = self.patient_data['DRE'].value_counts()
                suspicious_dre = sum(count for result, count in dre_counts.items() 
                                if result in ['Hard', 'Firm', 'Asymmetrical'])
                
                html += f"""
                <div class="metric-card">
                    <h4>Digital Rectal Exam Findings</h4>
                    <ul>
                        <li>Suspicious Findings: {suspicious_dre} patients ({suspicious_dre/total_patients:.1%})</li>
                """
                for result, count in dre_counts.items():
                    html += f"<li>{result}: {count} patients ({count/total_patients:.1%})</li>"
                html += "</ul></div>"

            # MRI Results
            if 'Mri_Result' in self.patient_data.columns:
                mri_counts = self.patient_data['Mri_Result'].value_counts()
                high_pirads = sum(count for result, count in mri_counts.items() 
                                if result in ['IV', 'V'])
                
                html += f"""
                <div class="metric-card">
                    <h4>MRI PIRADS Scores</h4>
                    <ul>
                        <li>High Risk (PIRADS 4-5): {high_pirads} patients ({high_pirads/total_patients:.1%})</li>
                """
                for result, count in mri_counts.items():
                    html += f"<li>PIRADS {result}: {count} patients ({count/total_patients:.1%})</li>"
                html += "</ul></div>"

            # Combined Risk Score Analysis
            html += """
            </div>  <!-- End executive-summary -->
            
            <div class="detailed-analysis">
                <h3>Detailed Risk Analysis</h3>
            """

            # Age-Specific Risk Profile
            if 'Age' in self.patient_data.columns and 'TPSA' in self.patient_data.columns:
                data = self.patient_data[['Age', 'TPSA']].dropna()
                if len(data) > 0:
                    data['AgeGroup'] = pd.cut(data['Age'], 
                                            bins=[0, 50, 60, 70, 80, 100],
                                            labels=['<50', '50-59', '60-69', '70-79', '≥80'])
                    
                    age_stats = data.groupby('AgeGroup')['TPSA'].agg(['mean', 'count', 'median'])
                    
                    html += """
                    <div class="age-analysis">
                        <h4>Age-Specific Risk Profile</h4>
                        <table class="data-table">
                            <tr>
                                <th>Age Group</th>
                                <th>Patients</th>
                                <th>Mean PSA</th>
                                <th>Median PSA</th>
                                <th>High PSA %</th>
                            </tr>
                    """
                    for group, stats in age_stats.iterrows():
                        high_psa_pct = len(data[(data['AgeGroup'] == group) & (data['TPSA'] > 4)]) / stats['count']
                        html += f"""
                        <tr>
                            <td>{group}</td>
                            <td>{int(stats['count'])}</td>
                            <td>{stats['mean']:.1f}</td>
                            <td>{stats['median']:.1f}</td>
                            <td>{high_psa_pct:.1%}</td>
                        </tr>
                        """
                    html += "</table></div>"

            # Risk Factor Correlations
            numeric_cols = ['Age', 'TPSA', 'FPSA', 'PSA', 'WG_Volume_cc', 'A_Volume_cc']
            available_cols = [col for col in numeric_cols if col in self.patient_data.columns]
            
            if len(available_cols) >= 2:
                corr_matrix = self.patient_data[available_cols].corr()
                
                html += """
                <div class="correlation-analysis">
                    <h4>Risk Factor Correlations</h4>
                    <table class="correlation-table">
                        <tr><th>Parameter</th>
                """
                html += "".join(f"<th>{col}</th>" for col in available_cols)
                html += "</tr>"
                
                for i, row in corr_matrix.iterrows():
                    html += f"<tr><td>{i}</td>"
                    for col in available_cols:
                        value = row[col]
                        color = "green" if abs(value) < 0.3 else ("orange" if abs(value) < 0.7 else "red")
                        html += f"<td style='background-color: {color}; color: white;'>{value:.2f}</td>"
                    html += "</tr>"
                html += "</table></div>"

            # Risk Score Distribution
            html += """
            <div class="risk-score-distribution">
                <h4>Risk Score Distribution</h4>
                <p>Calculated using PSA levels, DRE findings, and MRI results</p>
                <div class="risk-categories">
            """
            
            # Calculate risk scores (example implementation)
            if not self.patient_data.empty:
                risk_scores = []
                for _, patient in self.patient_data.iterrows():
                    score = 0
                    # PSA score
                    if 'TPSA' in patient and pd.notna(patient['TPSA']):
                        if patient['TPSA'] > 10: score += 3
                        elif patient['TPSA'] > 4: score += 2
                        else: score += 1
                    
                    # DRE score
                    if 'DRE' in patient and pd.notna(patient['DRE']):
                        if patient['DRE'] == 'Hard': score += 3
                        elif patient['DRE'] in ['Firm', 'Asymmetrical']: score += 2
                        else: score += 1
                    
                    # MRI score
                    if 'Mri_Result' in patient and pd.notna(patient['Mri_Result']):
                        if patient['Mri_Result'] in ['IV', 'V']: score += 3
                        elif patient['Mri_Result'] == 'III': score += 2
                    
                    risk_scores.append(score)
                
                if risk_scores:
                    risk_df = pd.DataFrame({'RiskScore': risk_scores})
                    risk_df['Category'] = pd.cut(risk_df['RiskScore'],
                                            bins=[0, 3, 6, 10],
                                            labels=['Low', 'Medium', 'High'])
                    risk_counts = risk_df['Category'].value_counts()
                    
                    html += "<table class='risk-table'><tr>"
                    for category, count in risk_counts.items():
                        pct = count / len(risk_scores)
                        html += f"""
                        <td class='risk-category {category.lower()}'>
                            <div class='risk-header'>{category} Risk</div>
                            <div class='risk-count'>{count}</div>
                            <div class='risk-pct'>{pct:.1%}</div>
                        </td>
                        """
                    html += "</tr></table>"
            
            html += """
                </div>
            </div>  <!-- End detailed-analysis -->
            
            <div class="recommendations">
                <h3>Clinical Recommendations</h3>
                <ul>
                    <li>Patients with High Risk scores should be prioritized for biopsy</li>
                    <li>Medium Risk patients may benefit from additional diagnostic tests</li>
                    <li>Low Risk patients should be monitored with regular PSA testing</li>
                </ul>
            </div>
        </div>  <!-- End risk-report -->
        
        <style>
            .risk-report {
                font-family: 'Segoe UI', Arial, sans-serif;
                max-width: 1200px;
                margin: 0 auto;
                color: #333;
            }
            .executive-summary {
                background-color: #f8f9fa;
                padding: 20px;
                border-radius: 8px;
                margin-bottom: 30px;
            }
            .metric-card {
                background: white;
                border-left: 4px solid #1e3a8a;
                padding: 15px;
                margin-bottom: 15px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            .data-table {
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }
            .data-table th, .data-table td {
                padding: 12px;
                text-align: center;
                border: 1px solid #ddd;
            }
            .data-table th {
                background-color: #1e3a8a;
                color: white;
            }
            .data-table tr:nth-child(even) {
                background-color: #f2f2f2;
            }
            .correlation-table {
                width: auto;
                margin: 20px 0;
                border-collapse: collapse;
            }
            .correlation-table th, .correlation-table td {
                padding: 10px;
                border: 1px solid #ddd;
                text-align: center;
            }
            .risk-table {
                width: 100%;
                margin: 20px 0;
                border-collapse: separate;
                border-spacing: 10px;
            }
            .risk-category {
                text-align: center;
                padding: 15px;
                border-radius: 8px;
                width: 33%;
            }
            .risk-category.low {
                background-color: #4ade80;
            }
            .risk-category.medium {
                background-color: #fbbf24;
            }
            .risk-category.high {
                background-color: #f87171;
            }
            .risk-header {
                font-weight: bold;
                font-size: 1.2em;
                margin-bottom: 5px;
            }
            .risk-count {
                font-size: 1.5em;
                font-weight: bold;
                margin: 5px 0;
            }
            .risk-pct {
                font-size: 1.1em;
            }
            .recommendations {
                background-color: #eef2ff;
                padding: 20px;
                border-radius: 8px;
                margin-top: 30px;
            }
            h2, h3, h4 {
                color: #1e3a8a;
            }
            h2 {
                border-bottom: 2px solid #1e3a8a;
                padding-bottom: 10px;
            }
            .report-date {
                color: #6b7280;
                text-align: right;
                font-style: italic;
            }
        </style>
        """
        
        return html
    def add_clinical_to_pdf(self, pdf):
        """Add clinical report content to PDF"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Clinical Statistics", ln=1)
        pdf.set_font("Arial", size=12)
        
        # Add PSA distribution if available
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="PSA Distribution", ln=1)
            fig = self.create_psa_distribution_figure()
            self.add_figure_to_pdf(pdf, "PSA Distribution", fig)
            plt.close(fig)
        
        # Add DRE results if available
        if not self.patient_data.empty and 'DRE' in self.patient_data.columns:
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="DRE Results", ln=1)
            
            dre_counts = self.patient_data['DRE'].value_counts()
            for result, count in dre_counts.items():
                pdf.cell(200, 10, txt=f"{result}: {count} patients", ln=1)
        
        # Add MRI results if available
        if not self.patient_data.empty and 'Mri_Result' in self.patient_data.columns:
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="MRI Results (PIRADS)", ln=1)
            
            mri_counts = self.patient_data['Mri_Result'].value_counts()
            for result, count in mri_counts.items():
                pdf.cell(200, 10, txt=f"PIRADS {result}: {count} patients", ln=1)
    def add_trends_to_pdf(self, pdf):
        """Add trends analysis content to PDF"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Trends Analysis", ln=1)
        pdf.set_font("Arial", size=12)
        
        # Add temporal analysis if date data exists
        if not self.patient_data.empty and 'Date' in self.patient_data.columns:
            try:
                # Create temporal trends figure
                dates = pd.to_datetime(self.patient_data['Date'], errors='coerce')
                dates = dates.dropna()
                
                if len(dates) > 0:
                    pdf.ln(10)
                    pdf.set_font("Arial", 'B', 12)
                    pdf.cell(200, 10, txt="Patient Registration Trends", ln=1)
                    
                    monthly_counts = dates.groupby(dates.dt.to_period('M')).size()
                    fig, ax = plt.subplots(figsize=(8, 4))
                    monthly_counts.plot(kind='line', ax=ax, color=self.colors['primary'])
                    ax.set_title('Patient Registration Over Time')
                    ax.grid(True, alpha=0.3)
                    
                    self.add_figure_to_pdf(pdf, "Monthly Patient Registration", fig)
                    plt.close(fig)
            except Exception as e:
                pdf.cell(200, 10, txt=f"Error generating trends: {str(e)}", ln=1)
        
        # Add PSA vs Age trend if data exists
        if (not self.patient_data.empty and 
            'Age' in self.patient_data.columns and 
            'TPSA' in self.patient_data.columns):
            
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="PSA vs Age Trend", ln=1)
            
            data = self.patient_data[['Age', 'TPSA']].dropna()
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.scatter(data['Age'], data['TPSA'], color=self.colors['accent'], alpha=0.6)
            
            # Add trend line if enough data points
            if len(data) > 2:
                z = np.polyfit(data['Age'], data['TPSA'], 1)
                p = np.poly1d(z)
                ax.plot(data['Age'], p(data['Age']), color=self.colors['danger'], linestyle='--')
                ax.set_title(f'PSA vs Age (Trend: y={z[0]:.2f}x + {z[1]:.2f})')
            else:
                ax.set_title('PSA vs Age')
                
            ax.grid(True, alpha=0.3)
            self.add_figure_to_pdf(pdf, "PSA vs Age Relationship", fig)
            plt.close(fig)
        
        # Add correlation matrix if enough numeric data exists
        numeric_cols = ['Age', 'TPSA', 'FPSA', 'PSA', 'WG_Volume_cc', 'A_Volume_cc']
        available_cols = [col for col in numeric_cols if col in self.patient_data.columns]
        
        if len(available_cols) >= 2:
            pdf.ln(10)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="Risk Factor Correlations", ln=1)
            
            corr_data = self.patient_data[available_cols].corr()
            fig, ax = plt.subplots(figsize=(8, 6))
            sns.heatmap(corr_data, annot=True, cmap='coolwarm', ax=ax)
            ax.set_title('Clinical Parameter Correlations')
            
            self.add_figure_to_pdf(pdf, "Correlation Matrix", fig)
            plt.close(fig)
    def add_risk_to_pdf(self, pdf):
        """Add risk assessment content to PDF"""
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Risk Assessment", ln=1)
        pdf.set_font("Arial", size=12)
        
        if not self.patient_data.empty:
            # Add basic risk metrics
            if 'TPSA' in self.patient_data.columns:
                high_psa = len(self.patient_data[self.patient_data['TPSA'] > 4])
                pdf.cell(200, 10, txt=f"Patients with PSA >4: {high_psa}", ln=1)
            
            if 'DRE' in self.patient_data.columns:
                suspicious_dre = len(self.patient_data[self.patient_data['DRE'].isin(['Hard', 'Firm'])])
                pdf.cell(200, 10, txt=f"Suspicious DRE findings: {suspicious_dre}", ln=1)
            
            # Add risk categories chart if possible
            if 'TPSA' in self.patient_data.columns:
                pdf.ln(10)
                psa_data = self.patient_data['TPSA'].dropna()
                low = len(psa_data[psa_data <= 4])
                moderate = len(psa_data[(psa_data > 4) & (psa_data <= 10)])
                high = len(psa_data[psa_data > 10])
                
                fig, ax = plt.subplots()
                ax.bar(['Low', 'Moderate', 'High'], [low, moderate, high], 
                    color=[self.colors['success'], self.colors['warning'], self.colors['danger']])
                ax.set_title('PSA Risk Categories')
                self.add_figure_to_pdf(pdf, "PSA Risk Stratification", fig)
                plt.close(fig)
    def _generate_summary_html(self) -> str:
        """
        Generate a clean HTML summary report with key patient statistics.
        Uses simple, reliable HTML formatting to avoid parsing issues.
        """
        # Initialize HTML content
        html_content = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            "<title>Patient Statistics Summary</title>",
            "<style>",
            "  body { font-family: Arial, sans-serif; margin: 20px; }",
            "  h2 { color: #2a5885; }",
            "  table { border-collapse: collapse; width: 100%; }",
            "  th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "  th { background-color: #0078d4; color: white; }",
            "  tr:nth-child(even) { background-color: #f2f2f2; }",
            "</style>",
            "</head>",
            "<body>",
            "<h2>Patient Statistics Summary</h2>",
            "<table>",
            "<tr><th>Metric</th><th>Value</th><th>Notes</th></tr>"
        ]

        try:
            # Basic patient count
            patient_count = len(self.patient_data)
            html_content.append(f"<tr><td>Total Patients</td><td>{patient_count}</td><td></td></tr>")

            # Age statistics
            if not self.patient_data.empty and 'Age' in self.patient_data.columns:
                age_stats = [
                    ("Average Age", f"{self.patient_data['Age'].mean():.1f}"),
                    ("Age Range", f"{self.patient_data['Age'].min()}-{self.patient_data['Age'].max()}")
                ]
                for metric, value in age_stats:
                    html_content.append(f"<tr><td>{metric}</td><td>{value}</td><td></td></tr>")

            # PSA statistics
            if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
                high_psa = self.patient_data['TPSA'] > 4
                high_psa_count = high_psa.sum()
                high_psa_pct = (high_psa_count / patient_count * 100) if patient_count > 0 else 0
                
                psa_stats = [
                    ("Patients with PSA >4", f"{high_psa_count} ({high_psa_pct:.1f}%)"),
                    ("Highest PSA", f"{self.patient_data['TPSA'].max():.1f}")
                ]
                for metric, value in psa_stats:
                    html_content.append(f"<tr><td>{metric}</td><td>{value}</td><td></td></tr>")

            # Family history
            if not self.patient_data.empty and 'Family' in self.patient_data.columns:
                family_history = (self.patient_data['Family'] == '+ve').sum()
                html_content.append(
                    f"<tr><td>Family History (+ve)</td><td>{family_history}</td><td></td></tr>"
                )

            # Add timestamp
            from datetime import datetime
            html_content.append(
                f"<tr><td>Report Generated</td><td colspan='2'>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td></tr>"
            )

        except Exception as e:
            html_content.append(
                f"<tr><td colspan='3' style='color: red;'>Error generating report: {str(e)}</td></tr>"
            )

        # Close HTML tags
        html_content.extend([
            "</table>",
            "</body>",
            "</html>"
        ])

        return "\n".join(html_content)
    def create_main_content(self):
        """Create main content area with tabbed interface"""
        self.main_frame = CTkFrame(self, fg_color=self.colors['light'])
        self.main_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        
        # Create tabview
        self.tabview = CTkTabview(self.main_frame, 
                                 fg_color=self.colors['white'],
                                 segmented_button_fg_color=self.colors['primary'],
                                 segmented_button_selected_color=self.colors['accent'])
        self.tabview.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Add tabs
        self.overview_tab = self.tabview.add("Overview")
        self.demographics_tab = self.tabview.add("Demographics")
        self.clinical_tab = self.tabview.add("Clinical")
        self.trends_tab = self.tabview.add("Trends")
        self.risk_tab = self.tabview.add("Risk Analysis")
        self.reports_tab = self.tabview.add("Reports")  # Add reports tab
        self.settings_tab = self.tabview.add("Settings")
        # Initialize with overview
        self.show_overview()
        
    def load_data(self):
        """Load and cache data from database"""
        try:
            self.conn = sqlite3.connect(self.db_path)
            
            # Initialize with empty DataFrames
            self.patient_data = pd.DataFrame()
            self.side_data = pd.DataFrame()
            self.bottles_data = pd.DataFrame()
            
            # Check if tables exist before trying to load them
            cursor = self.conn.cursor()
            
            # Load patient info if table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='PatientsInfo'")
            if cursor.fetchone():
                self.patient_data = pd.read_sql_query("SELECT * FROM PatientsInfo", self.conn)
            
            # Load side data if table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='PatientsSide'")
            if cursor.fetchone():
                self.side_data = pd.read_sql_query("SELECT * FROM PatientsSide", self.conn)
            
            # Load bottles data if table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='FirstBottles'")
            if cursor.fetchone():
                self.bottles_data = pd.read_sql_query("SELECT * FROM FirstBottles", self.conn)
            
            # Update status
            total_patients = len(self.patient_data)
            status_text = f"✅ {total_patients} patients loaded\n📊 Analysis ready"
            if hasattr(self, 'status_text'):
                self.status_text.configure(text=status_text)
            
        except Exception as e:
            error_text = f"❌ Error: {str(e)}"
            if hasattr(self, 'status_text'):
                self.status_text.configure(text=error_text)
            # Ensure we have empty DataFrames for error handling
            self.patient_data = pd.DataFrame()
            self.side_data = pd.DataFrame()
            self.bottles_data = pd.DataFrame()
            
    def show_overview(self):
        """Display comprehensive overview dashboard"""
        # Clear current content
        for widget in self.overview_tab.winfo_children():
            widget.destroy()
            
        # Create scrollable frame
        scroll_frame = CTkScrollableFrame(self.overview_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Key Metrics Cards Row
        metrics_frame = CTkFrame(scroll_frame, fg_color="transparent")
        metrics_frame.pack(fill="x", pady=(0, 20))
        metrics_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # Initialize default values
        total_patients = 0
        avg_age = 0
        family_history_positive = 0
        hard_dre_count = 0
        
        # Calculate key metrics if data exists
        if not self.patient_data.empty:
            total_patients = len(self.patient_data)
            if 'Age' in self.patient_data.columns:
                avg_age = self.patient_data['Age'].mean()
            if 'Family' in self.patient_data.columns:
                family_history_positive = len(self.patient_data[self.patient_data['Family'] == '+ve'])
            if 'DRE' in self.patient_data.columns:
                hard_dre_count = len(self.patient_data[self.patient_data['DRE'] == 'Hard'])
        
        # Rest of the method remains the same...
        # Create metric cards
        metrics = [
            ("Total Patients", total_patients, self.colors['primary'], "👥"),
            ("Average Age", f"{avg_age:.1f}", self.colors['accent'], "📊"),
            ("Family History +", family_history_positive, self.colors['warning'], "👨‍👩‍👧‍👦"),
            ("Hard DRE", hard_dre_count, self.colors['danger'], "🔍")
        ]
        
        for i, (title, value, color, icon) in enumerate(metrics):
            self.create_metric_card(metrics_frame, title, value, color, icon, i)
        
        # Charts Row
        charts_frame = CTkFrame(scroll_frame, fg_color="transparent")
        charts_frame.pack(fill="both", expand=True, pady=(0, 20))
        charts_frame.grid_columnconfigure((0, 1), weight=1)
        charts_frame.grid_rowconfigure(0, weight=1)
        
        # Age Distribution Chart
        age_chart_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        age_chart_frame.grid(row=0, column=0, padx=(0, 10), sticky="nsew")
        self.create_age_distribution_chart(age_chart_frame)
        
        # PSA Levels Chart
        psa_chart_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        psa_chart_frame.grid(row=0, column=1, padx=(10, 0), sticky="nsew")
        self.create_psa_distribution_chart(psa_chart_frame)
        
        # Recent Activity Section
        activity_frame = CTkFrame(scroll_frame, fg_color=self.colors['white'])
        activity_frame.pack(fill="x", pady=(0, 20))
        
        activity_title = CTkLabel(activity_frame, text="Recent Activity", 
                                 font=("Arial Black", 16, "bold"), 
                                 text_color=self.colors['primary'])
        activity_title.pack(pady=15)
        
        self.create_recent_activity_table(activity_frame)
        
    def create_metric_card(self, parent, title, value, color, icon, column):
        """Create a professional metric card"""
        card = CTkFrame(parent, fg_color=self.colors['white'], 
                       border_width=2, border_color=color, corner_radius=15)
        card.grid(row=0, column=column, padx=10, pady=10, sticky="ew")
        
        # Icon
        icon_label = CTkLabel(card, text=icon, font=("Arial", 24))
        icon_label.pack(pady=(15, 5))
        
        # Value
        value_label = CTkLabel(card, text=str(value), 
                              font=("Arial Black", 28, "bold"), text_color=color)
        value_label.pack()
        
        # Title
        title_label = CTkLabel(card, text=title, 
                              font=("Arial", 12), text_color=self.colors['text_dark'])
        title_label.pack(pady=(5, 15))
        
    def create_age_distribution_chart(self, parent):
        """Create age distribution histogram"""
        title_label = CTkLabel(parent, text="Age Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            # Filter out NaN values
            data = self.patient_data[['Age', 'TPSA']].dropna()
            
            scatter = ax.scatter(data['Age'], data['TPSA'], 
                               alpha=0.6, c=self.colors['primary'], s=50)
            ax.set_xlabel('Age (years)')
            ax.set_ylabel('TPSA (ng/ml)')
            ax.set_title('PSA vs Age Relationship')
            ax.grid(True, alpha=0.3)
            
            # Add trend line
            if len(data) > 1:
                z = np.polyfit(data['Age'], data['TPSA'], 1)
                p = np.poly1d(z)
                ax.plot(data['Age'].sort_values(), p(data['Age'].sort_values()), 
                       color=self.colors['danger'], linestyle='--', alpha=0.8)
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No PSA/Age data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_dre_distribution_chart(self, parent):
        """Create DRE results distribution chart"""
        title_label = CTkLabel(parent, text="DRE Results Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'DRE' in self.patient_data.columns:
            dre_counts = self.patient_data['DRE'].value_counts()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            # Color code based on DRE result severity
            colors = []
            for dre_type in dre_counts.index:
                if dre_type == 'Hard':
                    colors.append(self.colors['danger'])
                elif dre_type == 'Firm':
                    colors.append(self.colors['warning'])
                elif dre_type == 'Asymmetrical':
                    colors.append(self.colors['secondary'])
                else:  # Soft
                    colors.append(self.colors['success'])
            
            bars = ax.bar(dre_counts.index, dre_counts.values, color=colors)
            ax.set_xlabel('DRE Result')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Digital Rectal Examination Results')
            plt.xticks(rotation=45)
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No DRE data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_volume_distribution_chart(self, parent):
        """Create prostate volume distribution chart"""
        title_label = CTkLabel(parent, text="Prostate Volume Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'WG_Volume_cc' in self.patient_data.columns:
            volumes = self.patient_data['WG_Volume_cc'].dropna()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            ax.hist(volumes, bins=20, color=self.colors['accent'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('Volume (cc)')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Whole Gland Volume Distribution')
            ax.grid(True, alpha=0.3)
            
            # Add mean line
            mean_volume = volumes.mean()
            ax.axvline(mean_volume, color=self.colors['danger'], linestyle='--', 
                      label=f'Mean: {mean_volume:.1f}cc')
            ax.legend()
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No volume data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_mri_results_chart(self, parent):
        """Create MRI results distribution chart"""
        title_label = CTkLabel(parent, text="MRI Results (PIRADS)", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'Mri_Result' in self.patient_data.columns:
            mri_counts = self.patient_data['Mri_Result'].value_counts()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            # Color code based on PIRADS score
            colors = []
            for mri_result in mri_counts.index:
                if mri_result in ['IV', 'V']:
                    colors.append(self.colors['danger'])
                elif mri_result == 'III':
                    colors.append(self.colors['warning'])
                elif mri_result in ['I', 'II']:
                    colors.append(self.colors['success'])
                else:  # Not Done, PIRADS
                    colors.append(self.colors['text_light'])
            
            wedges, texts, autotexts = ax.pie(mri_counts.values, labels=mri_counts.index, 
                                             autopct='%1.1f%%', startangle=90, colors=colors)
            ax.set_title('MRI PIRADS Distribution')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No MRI data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def show_trends(self):
        """Display trends and patterns analysis"""
        # Clear current content
        for widget in self.trends_tab.winfo_children():
            widget.destroy()
            
        scroll_frame = CTkScrollableFrame(self.trends_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Trends header
        header_frame = CTkFrame(scroll_frame, fg_color=self.colors['secondary'])
        header_frame.pack(fill="x", pady=(0, 20))
        
        header_label = CTkLabel(header_frame, text="Trends & Patterns Analysis", 
                               font=("Arial Black", 18, "bold"), 
                               text_color=self.colors['white'])
        header_label.pack(pady=15)
        
        # Time-based analysis
        self.create_temporal_analysis(scroll_frame)
        
        # Pattern analysis
        self.create_pattern_analysis(scroll_frame)
        
    def create_temporal_analysis(self, parent):
        """Create temporal trends analysis"""
        temporal_frame = CTkFrame(parent, fg_color=self.colors['white'])
        temporal_frame.pack(fill="x", pady=(0, 20))
        
        title_label = CTkLabel(temporal_frame, text="Patient Registration Trends", 
                              font=("Arial Black", 16, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=15)
        
        if not self.patient_data.empty and 'Date' in self.patient_data.columns:
            try:
                # Parse dates and create time series
                dates = pd.to_datetime(self.patient_data['Date'], errors='coerce')
                dates = dates.dropna()
                
                if len(dates) > 0:
                    # Group by month
                    monthly_counts = dates.groupby(dates.dt.to_period('M')).size()
                    
                    fig, ax = plt.subplots(figsize=(12, 4))
                    fig.patch.set_facecolor('white')
                    
                    monthly_counts.plot(kind='line', ax=ax, color=self.colors['primary'], 
                                       marker='o', linewidth=2, markersize=6)
                    ax.set_xlabel('Month')
                    ax.set_ylabel('Number of Patients')
                    ax.set_title('Patient Registration Trend Over Time')
                    ax.grid(True, alpha=0.3)
                    plt.xticks(rotation=45)
                    
                    plt.tight_layout()
                    
                    canvas = FigureCanvasTkAgg(fig, temporal_frame)
                    canvas.draw()
                    canvas.get_tk_widget().pack(fill="both", expand=True, padx=15, pady=15)
                else:
                    no_data_label = CTkLabel(temporal_frame, text="No valid date data available", 
                                            text_color=self.colors['text_light'])
                    no_data_label.pack(expand=True)
            except Exception as e:
                error_label = CTkLabel(temporal_frame, text=f"Error processing dates: {str(e)}", 
                                      text_color=self.colors['danger'])
                error_label.pack(expand=True)
        else:
            no_data_label = CTkLabel(temporal_frame, text="No date data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_pattern_analysis(self, parent):
        """Create pattern analysis charts"""
        patterns_frame = CTkFrame(parent, fg_color="transparent")
        patterns_frame.pack(fill="both", expand=True)
        patterns_frame.grid_columnconfigure((0, 1), weight=1)
        patterns_frame.grid_rowconfigure(0, weight=1)
        
        # Age vs PSA pattern
        age_psa_pattern_frame = CTkFrame(patterns_frame, fg_color=self.colors['white'])
        age_psa_pattern_frame.grid(row=0, column=0, padx=(0, 10), sticky="nsew")
        self.create_age_psa_pattern_chart(age_psa_pattern_frame)
        
        # Risk factor correlation
        risk_correlation_frame = CTkFrame(patterns_frame, fg_color=self.colors['white'])
        risk_correlation_frame.grid(row=0, column=1, padx=(10, 0), sticky="nsew")
        self.create_risk_correlation_chart(risk_correlation_frame)
        
    def create_age_psa_pattern_chart(self, parent):
        """Create age vs PSA pattern analysis"""
        title_label = CTkLabel(parent, text="Age-PSA Pattern Analysis", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if (not self.patient_data.empty and 
            'Age' in self.patient_data.columns and 
            'TPSA' in self.patient_data.columns and 
            'DRE' in self.patient_data.columns):
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            # Create scatter plot colored by DRE result
            dre_colors = {
                'Hard': self.colors['danger'],
                'Firm': self.colors['warning'],
                'Asymmetrical': self.colors['secondary'],
                'Soft': self.colors['success']
            }
            
            for dre_type, color in dre_colors.items():
                mask = self.patient_data['DRE'] == dre_type
                data = self.patient_data[mask]
                if not data.empty:
                    ax.scatter(data['Age'], data['TPSA'], 
                             c=color, label=dre_type, alpha=0.7, s=50)
            
            ax.set_xlabel('Age (years)')
            ax.set_ylabel('TPSA (ng/ml)')
            ax.set_title('Age vs PSA by DRE Result')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="Insufficient data for pattern analysis", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_risk_correlation_chart(self, parent):
        """Create risk factor correlation heatmap"""
        title_label = CTkLabel(parent, text="Risk Factor Correlations", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty:
            # Select numeric columns for correlation
            numeric_cols = ['Age', 'TPSA', 'FPSA', 'PSA', 'WG_Volume_cc', 'A_Volume_cc']
            available_cols = [col for col in numeric_cols if col in self.patient_data.columns]
            
            if len(available_cols) >= 2:
                corr_data = self.patient_data[available_cols].corr()
                
                fig, ax = plt.subplots(figsize=(6, 4))
                fig.patch.set_facecolor('white')
                
                im = ax.imshow(corr_data.values, cmap='coolwarm', aspect='auto', vmin=-1, vmax=1)
                
                # Set ticks and labels
                ax.set_xticks(range(len(available_cols)))
                ax.set_yticks(range(len(available_cols)))
                ax.set_xticklabels(available_cols, rotation=45, ha='right')
                ax.set_yticklabels(available_cols)
                
                # Add correlation values
                for i in range(len(available_cols)):
                    for j in range(len(available_cols)):
                        text = ax.text(j, i, f'{corr_data.iloc[i, j]:.2f}',
                                     ha="center", va="center", color="black", fontsize=10)
                
                ax.set_title('Risk Factor Correlation Matrix')
                plt.colorbar(im, ax=ax)
                plt.tight_layout()
                
                canvas = FigureCanvasTkAgg(fig, parent)
                canvas.draw()
                canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
            else:
                no_data_label = CTkLabel(parent, text="Insufficient numeric data for correlation", 
                                        text_color=self.colors['text_light'])
                no_data_label.pack(expand=True)
        else:
            no_data_label = CTkLabel(parent, text="No data available for correlation analysis", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def show_risk_analysis(self):
        """Display risk assessment analysis"""
        # Clear current content
        for widget in self.risk_tab.winfo_children():
            widget.destroy()
            
        scroll_frame = CTkScrollableFrame(self.risk_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Risk analysis header
        header_frame = CTkFrame(scroll_frame, fg_color=self.colors['warning'])
        header_frame.pack(fill="x", pady=(0, 20))
        
        header_label = CTkLabel(header_frame, text="Risk Assessment & Stratification", 
                               font=("Arial Black", 18, "bold"), 
                               text_color=self.colors['white'])
        header_label.pack(pady=15)
        
        # Risk metrics
        self.create_risk_metrics(scroll_frame)
        
        # Risk stratification charts
        self.create_risk_charts(scroll_frame)
        
    def create_risk_metrics(self, parent):
        """Create risk assessment metrics"""
        metrics_frame = CTkFrame(parent, fg_color="transparent")
        metrics_frame.pack(fill="x", pady=(0, 20))
        metrics_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        if not self.patient_data.empty:
            # Calculate risk metrics
            high_risk_psa = len(self.patient_data[self.patient_data['TPSA'] > 10]) if 'TPSA' in self.patient_data.columns else 0
            suspicious_dre = len(self.patient_data[self.patient_data['DRE'].isin(['Hard', 'Asymmetrical'])]) if 'DRE' in self.patient_data.columns else 0
            high_pirads = len(self.patient_data[self.patient_data['Mri_Result'].isin(['IV', 'V'])]) if 'Mri_Result' in self.patient_data.columns else 0
            family_history_pos = len(self.patient_data[self.patient_data['Family'] == '+ve']) if 'Family' in self.patient_data.columns else 0
            
            risk_metrics = [
                ("High PSA (>10)", high_risk_psa, self.colors['danger'], "🚨"),
                ("Suspicious DRE", suspicious_dre, self.colors['warning'], "👆"),
                ("High PIRADS", high_pirads, self.colors['danger'], "📊"),
                ("Family History+", family_history_pos, self.colors['secondary'], "👨‍👩‍👧‍👦")
            ]
            
            for i, (title, value, color, icon) in enumerate(risk_metrics):
                self.create_metric_card(metrics_frame, title, value, color, icon, i)
                
    def create_risk_charts(self, parent):
        """Create risk stratification charts"""
        charts_frame = CTkFrame(parent, fg_color="transparent")
        charts_frame.pack(fill="both", expand=True)
        charts_frame.grid_columnconfigure((0, 1), weight=1)
        charts_frame.grid_rowconfigure((0, 1), weight=1)
        
        # PSA risk stratification
        psa_risk_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        psa_risk_frame.grid(row=0, column=0, padx=(0, 10), pady=(0, 10), sticky="nsew")
        self.create_psa_risk_chart(psa_risk_frame)
        
        # Combined risk assessment
        combined_risk_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        combined_risk_frame.grid(row=0, column=1, padx=(10, 0), pady=(0, 10), sticky="nsew")
        self.create_combined_risk_chart(combined_risk_frame)
        
        # Age-specific risk
        age_risk_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        age_risk_frame.grid(row=1, column=0, padx=(0, 10), pady=(10, 0), sticky="nsew")
        self.create_age_risk_chart(age_risk_frame)
        
        # Risk progression
        progression_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        progression_frame.grid(row=1, column=1, padx=(10, 0), pady=(10, 0), sticky="nsew")
        self.create_risk_progression_chart(progression_frame)
        
    def create_psa_risk_chart(self, parent):
        """Create PSA risk stratification chart"""
        title_label = CTkLabel(parent, text="PSA Risk Stratification", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            psa_data = self.patient_data['TPSA'].dropna()
            
            # Define PSA risk categories
            low_risk = len(psa_data[psa_data <= 4])
            moderate_risk = len(psa_data[(psa_data > 4) & (psa_data <= 10)])
            high_risk = len(psa_data[psa_data > 10])
            
            categories = ['Low Risk\n(≤4)', 'Moderate Risk\n(4-10)', 'High Risk\n(>10)']
            values = [low_risk, moderate_risk, high_risk]
            colors = [self.colors['success'], self.colors['warning'], self.colors['danger']]
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            bars = ax.bar(categories, values, color=colors)
            ax.set_ylabel('Number of Patients')
            ax.set_title('PSA Risk Categories')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No PSA data for risk analysis", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_combined_risk_chart(self, parent):
        """Create combined risk factor assessment"""
        title_label = CTkLabel(parent, text="Combined Risk Assessment", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty:
            # Calculate combined risk scores
            risk_scores = []
            
            for _, patient in self.patient_data.iterrows():
                score = 0
                
                # PSA risk
                if 'TPSA' in patient and pd.notna(patient['TPSA']):
                    if patient['TPSA'] > 10:
                        score += 3
                    elif patient['TPSA'] > 4:
                        score += 2
                    else:
                        score += 1
                
                # DRE risk
                if 'DRE' in patient and pd.notna(patient['DRE']):
                    if patient['DRE'] == 'Hard':
                        score += 3
                    elif patient['DRE'] in ['Firm', 'Asymmetrical']:
                        score += 2
                    else:
                        score += 1
                
                # Family history
                if 'Family' in patient and patient['Family'] == '+ve':
                    score += 2
                
                # MRI results
                if 'Mri_Result' in patient and pd.notna(patient['Mri_Result']):
                    if patient['Mri_Result'] in ['IV', 'V']:
                        score += 3
                    elif patient['Mri_Result'] == 'III':
                        score += 2
                
                risk_scores.append(score)
            
            if risk_scores:
                # Categorize risk scores
                low_risk = len([s for s in risk_scores if s <= 4])
                moderate_risk = len([s for s in risk_scores if 5 <= s <= 7])
                high_risk = len([s for s in risk_scores if s >= 8])
                
                fig, ax = plt.subplots(figsize=(6, 4))
                fig.patch.set_facecolor('white')
                
                sizes = [low_risk, moderate_risk, high_risk]
                labels = ['Low Risk', 'Moderate Risk', 'High Risk']
                colors = [self.colors['success'], self.colors['warning'], self.colors['danger']]
                
                wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%', 
                                                 colors=colors, startangle=90)
                ax.set_title('Combined Risk Distribution')
                
                plt.tight_layout()
                
                canvas = FigureCanvasTkAgg(fig, parent)
                canvas.draw()
                canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
            else:
                no_data_label = CTkLabel(parent, text="Insufficient data for risk calculation", 
                                        text_color=self.colors['text_light'])
                no_data_label.pack(expand=True)
        else:
            no_data_label = CTkLabel(parent, text="No data for combined risk analysis", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_age_risk_chart(self, parent):
        """Create age-specific risk analysis"""
        title_label = CTkLabel(parent, text="Age-Specific Risk Profile", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if (not self.patient_data.empty and 
            'Age' in self.patient_data.columns and 
            'TPSA' in self.patient_data.columns):
            
            # Create age groups
            data = self.patient_data[['Age', 'TPSA']].dropna()
            data['AgeGroup'] = pd.cut(data['Age'], bins=[0, 50, 60, 70, 80, 100], 
                                     labels=['<50', '50-59', '60-69', '70-79', '≥80'])
            
            # Calculate mean PSA by age group
            age_psa_means = data.groupby('AgeGroup')['TPSA'].mean()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            bars = ax.bar(age_psa_means.index.astype(str), age_psa_means.values, 
                         color=self.colors['secondary'])
            ax.set_xlabel('Age Group')
            ax.set_ylabel('Mean TPSA (ng/ml)')
            ax.set_title('Average PSA by Age Group')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{height:.1f}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No age/PSA data for analysis", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_risk_progression_chart(self, parent):
        """Create risk progression over time"""
        title_label = CTkLabel(parent, text="Risk Progression Timeline", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if (not self.patient_data.empty and 
            'Date' in self.patient_data.columns and 
            'TPSA' in self.patient_data.columns):
            
            try:
                # Parse dates and create timeline
                data = self.patient_data[['Date', 'TPSA']].copy()
                data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
                data = data.dropna()
                
                if len(data) > 0:
                    # Group by month and calculate high-risk percentage
                    data['Month'] = data['Date'].dt.to_period('M')
                    monthly_data = data.groupby('Month').agg({
                        'TPSA': ['count', lambda x: (x > 4).sum()]
                    }).reset_index()
                    
                    monthly_data.columns = ['Month', 'Total', 'HighRisk']
                    monthly_data['RiskPercentage'] = (monthly_data['HighRisk'] / monthly_data['Total']) * 100
                    
                    fig, ax = plt.subplots(figsize=(6, 4))
                    fig.patch.set_facecolor('white')
                    
                    ax.plot(range(len(monthly_data)), monthly_data['RiskPercentage'], 
                           color=self.colors['danger'], marker='o', linewidth=2)
                ax.set_facecolor('white')
            except Exception as e:
                showerror("Error", f"Failed to create risk progression chart: {str(e)}")
                return
            
            ax.set_xlabel('Month')
            ax.set_ylabel('Risk Percentage')
            ax.set_title('Risk Progression Over Time')
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No date/PSA data for analysis", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_age_distribution_chart(self, parent):
        """Create age distribution chart"""
        title_label = CTkLabel(parent, text="Patient Age Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if (not self.patient_data.empty and 
            'Age' in self.patient_data.columns):
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            ages = self.patient_data['Age'].dropna()
            ax.hist(ages, bins=15, color=self.colors['primary'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('Age')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Patient Age Distribution')
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No age data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_psa_distribution_chart(self, parent):
        """Create PSA levels distribution chart"""
        title_label = CTkLabel(parent, text="PSA Levels Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'TPSA' in self.patient_data.columns:
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            psa_data = self.patient_data['TPSA'].dropna()
            ax.hist(psa_data, bins=20, color=self.colors['accent'], alpha=0.7, edgecolor='white')
            ax.set_xlabel('TPSA (ng/ml)')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Total PSA Distribution')
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No PSA data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_recent_activity_table(self, parent):
        """Create recent activity table"""
        if not self.patient_data.empty:
            # Get recent patients (assuming Date column exists)
            recent_data = self.patient_data.head(10)  # Last 10 entries
            
            # Create treeview for table
            columns = ['Name', 'Age', 'Date', 'DRE', 'TPSA']
            tree = ttk.Treeview(parent, columns=columns, show='headings', height=8)
            
            # Configure columns
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor='center')
            
            # Insert data
            for _, row in recent_data.iterrows():
                values = []
                for col in columns:
                    if col in row:
                        values.append(str(row[col]) if pd.notna(row[col]) else 'N/A')
                    else:
                        values.append('N/A')
                tree.insert('', 'end', values=values)
            
            tree.pack(fill="both", expand=True, padx=15, pady=15)
        else:
            no_data_label = CTkLabel(parent, text="No recent activity data", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def show_demographics(self):
        """Display patient demographics analysis"""
        # Clear current content
        for widget in self.demographics_tab.winfo_children():
            widget.destroy()
            
        scroll_frame = CTkScrollableFrame(self.demographics_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Demographics header
        header_frame = CTkFrame(scroll_frame, fg_color=self.colors['primary'])
        header_frame.pack(fill="x", pady=(0, 20))
        
        header_label = CTkLabel(header_frame, text="Patient Demographics Analysis", 
                               font=("Arial Black", 18, "bold"), 
                               text_color=self.colors['white'])
        header_label.pack(pady=15)
        
        # Charts grid
        charts_grid = CTkFrame(scroll_frame, fg_color="transparent")
        charts_grid.pack(fill="both", expand=True)
        charts_grid.grid_columnconfigure((0, 1), weight=1)
        charts_grid.grid_rowconfigure((0, 1), weight=1)
        
        # Age groups pie chart
        age_groups_frame = CTkFrame(charts_grid, fg_color=self.colors['white'])
        age_groups_frame.grid(row=0, column=0, padx=(0, 10), pady=(0, 10), sticky="nsew")
        self.create_age_groups_chart(age_groups_frame)
        
        # Nationality distribution
        nationality_frame = CTkFrame(charts_grid, fg_color=self.colors['white'])
        nationality_frame.grid(row=0, column=1, padx=(10, 0), pady=(0, 10), sticky="nsew")
        self.create_nationality_chart(nationality_frame)
        
        # Family history analysis
        family_frame = CTkFrame(charts_grid, fg_color=self.colors['white'])
        family_frame.grid(row=1, column=0, padx=(0, 10), pady=(10, 0), sticky="nsew")
        self.create_family_history_chart(family_frame)
        
        # Hospital distribution
        hospital_frame = CTkFrame(charts_grid, fg_color=self.colors['white'])
        hospital_frame.grid(row=1, column=1, padx=(10, 0), pady=(10, 0), sticky="nsew")
        self.create_hospital_distribution_chart(hospital_frame)
        
    def create_age_groups_chart(self, parent):
        """Create age groups pie chart"""
        title_label = CTkLabel(parent, text="Age Groups Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'Age' in self.patient_data.columns:
            ages = self.patient_data['Age'].dropna()
            
            # Define age groups
            age_groups = pd.cut(ages, bins=[0, 50, 60, 70, 80, 100], 
                               labels=['<50', '50-59', '60-69', '70-79', '≥80'])
            age_counts = age_groups.value_counts()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            wedges, texts, autotexts = ax.pie(age_counts.values, labels=age_counts.index, 
                                             autopct='%1.1f%%', startangle=90,
                                             colors=self.colors['chart_colors'])
            ax.set_title('Age Groups Distribution')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No age data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_nationality_chart(self, parent):
        """Create nationality distribution chart"""
        title_label = CTkLabel(parent, text="Nationality Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'Nationality' in self.patient_data.columns:
            nationality_counts = self.patient_data['Nationality'].value_counts()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            bars = ax.bar(nationality_counts.index, nationality_counts.values, 
                         color=self.colors['chart_colors'][:len(nationality_counts)])
            ax.set_xlabel('Nationality')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Patient Nationality Distribution')
            plt.xticks(rotation=45)
            
            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No nationality data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_family_history_chart(self, parent):
        """Create family history analysis chart"""
        title_label = CTkLabel(parent, text="Family History Analysis", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'Family' in self.patient_data.columns:
            family_counts = self.patient_data['Family'].value_counts()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            colors = [self.colors['success'] if x == '+ve' else self.colors['primary'] 
                     for x in family_counts.index]
            
            wedges, texts, autotexts = ax.pie(family_counts.values, labels=family_counts.index, 
                                             autopct='%1.1f%%', startangle=90, colors=colors)
            ax.set_title('Family History Distribution')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No family history data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def create_hospital_distribution_chart(self, parent):
        """Create hospital distribution chart"""
        title_label = CTkLabel(parent, text="Hospital Distribution", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if not self.patient_data.empty and 'Hospital' in self.patient_data.columns:
            hospital_counts = self.patient_data['Hospital'].value_counts()
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor('white')
            
            bars = ax.bar(hospital_counts.index, hospital_counts.values, 
                         color=self.colors['chart_colors'][:len(hospital_counts)])
            ax.set_xlabel('Hospital')
            ax.set_ylabel('Number of Patients')
            ax.set_title('Patient Hospital Distribution')
            plt.xticks(rotation=45)
            
            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No hospital data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)
            
    def show_clinical(self):
        """Display clinical analysis"""
        # Clear current content
        for widget in self.clinical_tab.winfo_children():
            widget.destroy()
            
        scroll_frame = CTkScrollableFrame(self.clinical_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Clinical header
        header_frame = CTkFrame(scroll_frame, fg_color=self.colors['accent'])
        header_frame.pack(fill="x", pady=(0, 20))
        
        header_label = CTkLabel(header_frame, text="Clinical Data Analysis", 
                               font=("Arial Black", 18, "bold"), 
                               text_color=self.colors['white'])
        header_label.pack(pady=15)
        
        # Clinical metrics grid
        self.create_clinical_metrics(scroll_frame)
        
        # Clinical charts
        self.create_clinical_charts(scroll_frame)
        
    def create_clinical_metrics(self, parent):
        """Create clinical metrics cards"""
        metrics_frame = CTkFrame(parent, fg_color="transparent")
        metrics_frame.pack(fill="x", pady=(0, 20))
        metrics_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        if not self.patient_data.empty:
            # Calculate clinical metrics
            avg_psa = self.patient_data['TPSA'].mean() if 'TPSA' in self.patient_data.columns else 0
            high_psa_count = len(self.patient_data[self.patient_data['TPSA'] > 4]) if 'TPSA' in self.patient_data.columns else 0
            avg_volume = self.patient_data['WG_Volume_cc'].mean() if 'WG_Volume_cc' in self.patient_data.columns else 0
            mri_done = len(self.patient_data[self.patient_data['Mri_Result'] != 'Not Done']) if 'Mri_Result' in self.patient_data.columns else 0
            
            clinical_metrics = [
                ("Avg PSA", f"{avg_psa:.2f}", self.colors['primary'], "🧪"),
                ("High PSA (>4)", high_psa_count, self.colors['warning'], "⚠️"),
                ("Avg Volume", f"{avg_volume:.1f}cc", self.colors['accent'], "📏"),
                ("MRI Done", mri_done, self.colors['success'], "🏥")
            ]
            
            for i, (title, value, color, icon) in enumerate(clinical_metrics):
                self.create_metric_card(metrics_frame, title, value, color, icon, i)
                
    def create_clinical_charts(self, parent):
        """Create clinical analysis charts"""
        charts_frame = CTkFrame(parent, fg_color="transparent")
        charts_frame.pack(fill="both", expand=True)
        charts_frame.grid_columnconfigure((0, 1), weight=1)
        charts_frame.grid_rowconfigure((0, 1), weight=1)
        
        # PSA vs Age scatter plot
        psa_age_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        psa_age_frame.grid(row=0, column=0, padx=(0, 10), pady=(0, 10), sticky="nsew")
        self.create_psa_age_scatter(psa_age_frame)
        
        # DRE results distribution
        dre_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        dre_frame.grid(row=0, column=1, padx=(10, 0), pady=(0, 10), sticky="nsew")
        self.create_dre_distribution_chart(dre_frame)
        
        # Volume distribution
        volume_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        volume_frame.grid(row=1, column=0, padx=(0, 10), pady=(10, 0), sticky="nsew")
        self.create_volume_distribution_chart(volume_frame)
        
        # MRI results
        mri_frame = CTkFrame(charts_frame, fg_color=self.colors['white'])
        mri_frame.grid(row=1, column=1, padx=(10, 0), pady=(10, 0), sticky="nsew")
        self.create_mri_results_chart(mri_frame)
        
    def create_psa_age_scatter(self, parent):
        """Create PSA vs Age scatter plot"""
        title_label = CTkLabel(parent, text="PSA vs Age Correlation", 
                              font=("Arial Black", 14, "bold"), 
                              text_color=self.colors['primary'])
        title_label.pack(pady=10)
        
        if (not self.patient_data.empty and 
            'TPSA' in self.patient_data.columns and 
            'Age' in self.patient_data.columns):
            
            fig, ax = plt.subplots(figsize=(6, 4))
            fig.patch.set_facecolor(self.colors['white'])
            ax.scatter(self.patient_data['Age'], self.patient_data['TPSA'],
                       color=self.colors['accent'], alpha=0.7)
            ax.set_xlabel('Age (years)')
            ax.set_ylabel('PSA (ng/ml)')
            ax.set_title('PSA vs Age')
            plt.tight_layout()
            # Add trend line if there's enough data
            if len(self.patient_data) > 1:
                x = self.patient_data['Age'].dropna()
                y = self.patient_data['TPSA'].dropna()
                if len(x) > 1 and len(y) > 1:
                    z = np.polyfit(x, y, 1)
                    p = np.poly1d(z)
                    ax.plot(x, p(x), color=self.colors['danger'], linestyle='--', 
                           label=f'Trend: y={z[0]:.2f}x + {z[1]:.2f}')
                    ax.legend()
            
            canvas = FigureCanvasTkAgg(fig, parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        else:
            no_data_label = CTkLabel(parent, text="No PSA/Age data available", 
                                    text_color=self.colors['text_light'])
            no_data_label.pack(expand=True)

    def show_settings(self):
        """Display settings panel"""
        # Clear current content
        for widget in self.settings_tab.winfo_children():
            widget.destroy()
            
        scroll_frame = CTkScrollableFrame(self.settings_tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Settings header
        header_frame = CTkFrame(scroll_frame, fg_color=self.colors['dark'])
        header_frame.pack(fill="x", pady=(0, 20))
        
        header_label = CTkLabel(header_frame, text="Application Settings", 
                               font=("Arial Black", 18, "bold"), 
                               text_color=self.colors['white'])
        header_label.pack(pady=15)
        
        # Database settings
        db_frame = CTkFrame(scroll_frame, fg_color=self.colors['white'])
        db_frame.pack(fill="x", pady=(0, 20))
        
        db_title = CTkLabel(db_frame, text="Database Configuration", 
                           font=("Arial Black", 14, "bold"), 
                           text_color=self.colors['primary'])
        db_title.pack(pady=(15, 5), anchor="w", padx=15)
        
        # Database path
        db_path_frame = CTkFrame(db_frame, fg_color="transparent")
        db_path_frame.pack(fill="x", padx=15, pady=5)
        
        db_path_label = CTkLabel(db_path_frame, text="Database Path:", 
                                font=("Arial", 12), 
                                text_color=self.colors['text_dark'])
        db_path_label.pack(side="left")
        
        self.db_path_entry = ctk.CTkEntry(db_path_frame, 
                                         placeholder_text=self.db_path,
                                         width=300)
        self.db_path_entry.pack(side="left", padx=10)
        
        db_path_button = CTkButton(db_path_frame, text="Browse...",
                                  width=80, command=self.browse_db_path)
        db_path_button.pack(side="left")
        
        # Backup settings
        backup_frame = CTkFrame(db_frame, fg_color="transparent")
        backup_frame.pack(fill="x", padx=15, pady=5)
        
        backup_label = CTkLabel(backup_frame, text="Auto Backup:", 
                               font=("Arial", 12), 
                               text_color=self.colors['text_dark'])
        backup_label.pack(side="left")
        
        self.backup_var = ctk.StringVar(value="on")
        backup_switch = ctk.CTkSwitch(backup_frame, text="Enabled",
                                     variable=self.backup_var,
                                     onvalue="on", offvalue="off")
        backup_switch.pack(side="left", padx=10)
        
        # Save button
        save_button = CTkButton(db_frame, text="Save Database Settings",
                               fg_color=self.colors['accent'],
                               command=self.save_db_settings)
        save_button.pack(pady=15)
        
        # Appearance settings
        appearance_frame = CTkFrame(scroll_frame, fg_color=self.colors['white'])
        appearance_frame.pack(fill="x", pady=(0, 20))
        
        appearance_title = CTkLabel(appearance_frame, text="Appearance", 
                                   font=("Arial Black", 14, "bold"), 
                                   text_color=self.colors['primary'])
        appearance_title.pack(pady=(15, 5), anchor="w", padx=15)
        
        # Theme selection
        theme_frame = CTkFrame(appearance_frame, fg_color="transparent")
        theme_frame.pack(fill="x", padx=15, pady=5)
        
        theme_label = CTkLabel(theme_frame, text="Color Theme:", 
                               font=("Arial", 12), 
                               text_color=self.colors['text_dark'])
        theme_label.pack(side="left")
        
        self.theme_var = ctk.StringVar(value="dark")
        theme_options = ["dark", "light", "system"]
        theme_menu = ctk.CTkOptionMenu(theme_frame, values=theme_options,
                                      variable=self.theme_var,
                                      command=self.change_theme)
        theme_menu.pack(side="left", padx=10)
        
        # Chart settings
        chart_frame = CTkFrame(appearance_frame, fg_color="transparent")
        chart_frame.pack(fill="x", padx=15, pady=5)
        
        chart_label = CTkLabel(chart_frame, text="Chart Style:", 
                              font=("Arial", 12), 
                              text_color=self.colors['text_dark'])
        chart_label.pack(side="left")
        
        self.chart_style_var = ctk.StringVar(value="seaborn")
        chart_options = ["seaborn", "ggplot", "dark_background", "bmh"]
        chart_menu = ctk.CTkOptionMenu(chart_frame, values=chart_options,
                                      variable=self.chart_style_var,
                                      command=self.change_chart_style)
        chart_menu.pack(side="left", padx=10)
        
    def browse_db_path(self):
        """Open file dialog to browse for database path"""
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            title="Select Database File",
            filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")]
        )
        if file_path:
            self.db_path_entry.delete(0, "end")
            self.db_path_entry.insert(0, file_path)
            
    def save_db_settings(self):
        """Save database configuration settings"""
        new_path = self.db_path_entry.get()
        if new_path and new_path != self.db_path:
            self.db_path = new_path
            self.load_data()  # Reload data with new path
            
        backup_enabled = self.backup_var.get() == "on"
        # TODO: Implement backup functionality
        
        showinfo(title="Settings Saved", message="Database settings have been saved successfully")
        
    def change_theme(self, choice):
        """Change application theme"""
        ctk.set_appearance_mode(choice)
        
    def change_chart_style(self, choice):
        """Change matplotlib chart style"""
        plt.style.use(choice)
        self.update_all_charts()
        
    def update_all_charts(self):
        """Refresh all charts in the application"""
        # Store current tab
        current_tab = self.tabview.get()
        
        # Refresh each tab
        tabs = ["Overview", "Demographics", "Clinical", "Trends", "Risk Analysis"]
        for tab in tabs:
            self.tabview.set(tab)
            self.tabview.update()
            
        # Restore original tab
        self.tabview.set(current_tab)
    def create_biopsy_analysis(self):
        """Create biopsy results analysis (to be implemented)"""
        pass
        
    def create_treatment_outcomes(self):
        """Create treatment outcomes analysis (to be implemented)"""
        pass
    def __del__(self):
        """Clean up resources"""
        if hasattr(self, 'conn'):
            self.conn.close()